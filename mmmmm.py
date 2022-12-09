from bs4 import BeautifulSoup
import requests
import random
import openpyxl
from tkinter import *
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from time import sleep
import time
from selenium.webdriver.common.keys import Keys
import pandas as pd
from selenium.webdriver.firefox.options import Options
import threading
1

def btn_click():

    #excel= openpyxl.Workbook() 
    #print(excel.sheetnames)
    #sheet = excel.active
    #sheet.title= 'TopAnime'
    #print(excel.sheetnames)
    #sheet.append(['Номер', 'Название', 'Оценка'])
    options = webdriver.FirefoxOptions()
    options.set_preference('general.useragent.override','Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148')



    try:

        url = entry.get()
        user_agents_list = [
        'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.83 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
        ]   
        time.sleep(1)
        
        source = requests.get(url, headers={'User-Agent': random.choice(user_agents_list)})
        #with open("index.html", "w") as file:
        #    file.write(source.text)

        driver = webdriver.Firefox()
        options = options
        driver.get('https://shikimori.one/Mr.Incognito!/list/anime/mylist/completed')
        time.sleep(5)
        with open('index.selenium.html', 'w') as file:
            file.write(driver.page_source)

        with open('index.selenium.html') as file:
            src = file.read()

        source.raise_for_status()

        soup = BeautifulSoup(src,'html')
        
        animes = soup.find('tbody', class_="entries").find_all("tr")
        for anime in animes:
        
            index__ = anime.find("td", class_="index").span.text
            name = anime.find("td", class_="name").find("span", class_="name-ru").text       
            grade = anime.find("td", class_="num").text
            
            print(index__, name, grade)
            #sheet.append([index__, name, grade])
                
    except Exception as e:
        print(e)

    #excel.save('myanimelist.xlsx')


root = Tk()

root["bg"]='white'
root.title('XXXX')
root.wm_attributes('-alpha', 1)
root.geometry('400x100')

root.resizable(width=False, height=False)

entry = ttk.Entry()
entry.pack(anchor=NW, padx=6, pady=6)


btn = Button(root, text='da', command=btn_click)
btn.pack()

root.mainloop()


